# journalist_twitter_analysis.py

import os
import tweepy
from collections import Counter, defaultdict
import pytz
import pandas as pd
from pymongo import MongoClient
import unicodedata
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import yagmail
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

# ==== Secrets from GitHub Environment ====
MONGO_URI = os.getenv("MONGO_URI")
BEARER_TOKEN = os.getenv("TWITTER_BEARER")
SENDER_EMAIL = os.getenv("SENDER_EMAIL")
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD")
TO_EMAIL = os.getenv("TO_EMAIL")
CC_EMAIL = os.getenv("CC_EMAIL")

for var, name in [
    (MONGO_URI, "MONGO_URI"),
    (BEARER_TOKEN, "TWITTER_BEARER"),
    (SENDER_EMAIL, "SENDER_EMAIL"),
    (SENDER_PASSWORD, "SENDER_PASSWORD"),
    (TO_EMAIL, "TO_EMAIL"),
    (CC_EMAIL, "CC_EMAIL"),
]:
    if not var:
        raise ValueError(f"❌ Environment variable '{name}' is not set.")

# ==== MongoDB Setup ====
mongo_client = MongoClient(MONGO_URI)
db = mongo_client["twitter_analysis"]
collection = db["journalist_reports"]

# ==== Twitter Client ====
client = tweepy.Client(bearer_token=BEARER_TOKEN, wait_on_rate_limit=True)

# ==== Journalist Handles ====
journalist_handles = [
     "swachhhyd", "CoreenaSuares2", "Vasudha156", "payalmehta100",
    "TheNaveena", "jsuryareddy", "umasudhir", "SriLakshmi_10", "BSNMalleswarRao",
    "Avinashpujari02", "nabilajamal_", "YakkatiSowmith", "Hindu_vs", "UjwalB_Journo",
    "NageshT93116498", "crime_kumar", "KanizaGarari", "KP_Aashish", "balaexpressTNIE"
]
# ==== Party Keyword Dictionary ====
party_keywords = {
    "TDP": ["chandrababuNaidu", "ncbn", "tdp", "Ministernaralokesh", "naralokesh", "nandamuribalakrishna",
            "#cmchandrababu", "#tdp", "#naralokesh", "Narachandrababunaidu",
            "Vangalapudianitha", "@anitha_TDP",
            "చంద్రబాబు", "సీఎం చంద్రబాబు", "నారా లోకేష్", "లోకేష్", "బాలకృష్ణ", "టిడిపి",
            "మంత్రి లోకేష్", "మంత్రి నారా లోకేష్", "లోకేశ్"],
    "YCP": ["jagan", "ysjagan", "ysjaganmohanreddy", "ysr", "ysrcp", "#ysjagan", "#ycp", "#ysrcp", "vidadalarajini","ysVijayamma","#sajjalaramakrishnareddy","#botsasatyanarayana"],
    "JSP": ["pawankalyan", "janasena", "DeputyCMPawanKalyan"],
    "BJP": ["bjp", "modi", "amitshah", "narendra modi", "#bjp", "pmmodi"],
    "INC": ["rahul gandhi", "congress", "indian national congress", "yssharmila", "inc"]
}

# ==== Exact Keyword Mentions ====
specific_keywords = [
    "cmchandrababu", "ysjagan", "pawankalyan", "DeputyCMPawanKalyan",
    "tdp", "ysrcp", "naralokesh", "janasena",
    "pithapuram", "thallikivandanam", "rapparappa",
    "ncbn", "chandrababuNaidu", "చంద్రబాబు", "సీమ్ చంద్రబాబు",
    "నారా లోకేష్", "లోకేష్"
]
# ==== Government and Telangana Keywords ====
govt_keywords = [
    "government", "govt", "cabinet", "minister", "mla", "mp","apliquorscam"
    "policy", "scheme", "budget", "official", "administration","#apliquorscam"
    "ఆదేశాలు", "ప్రభుత్వం", "మంత్రివర్గం", "తీర్మానం"
]

telangana_keywords = [
    "hyderabad", "kcr", "ktr", "b.r.s", "brs", "telangana",
    "తెలంగాణ", "కేసీఆర్", "కేటీఆర్","cmrevanthreddy","INC"
]

# ==== Time Slot Configuration ====
time_slots = {
    "12 AM–2:59 AM": (0, 3),
    "3 AM–5:59 AM": (3, 6),
    "6 AM–8:59 AM": (6, 9),
    "9 AM–11:59 AM": (9, 12),
    "12 PM–2:59 PM": (12, 15),
    "3 PM–5:59 PM": (15, 18),
    "6 PM–8:59 PM": (18, 21),
    "9 PM–11:59 PM": (21, 24)
}

def get_time_slot(dt):
    for slot, (s, e) in time_slots.items():
        if s <= dt.hour < e:
            return slot
    return "Unknown"

ist = pytz.timezone("Asia/Kolkata")
target_date = dt.datetime.now(ist).date()
start_ist = dt.datetime.combine(target_date, dt.time(0, 0, tzinfo=ist))
end_ist = dt.datetime.combine(target_date, dt.time(23, 59, 59, tzinfo=ist))
start_time = start_ist.astimezone(pytz.UTC).isoformat()
end_time = end_ist.astimezone(pytz.UTC).isoformat()

def fetch_tweets(username, start_time, end_time, max_results=100):
    tweets = []
    try:
        user = client.get_user(username=username)
        if not user.data:
            return []
        uid = user.data.id
        paginator = tweepy.Paginator(
            client.get_users_tweets,
            id=uid,
            start_time=start_time,
            end_time=end_time,
            tweet_fields=["created_at", "public_metrics", "entities", "text"],
            max_results=max_results
        )
        for page in paginator:
            if page.data:
                tweets.extend(page.data)
    except Exception as e:
        print(f"Error fetching tweets for {username}: {e}")
    return tweets

def process_handle(handle):
    print(f"\n📥 Processing @{handle}...")
    tweets = fetch_tweets(handle, start_time, end_time)
    
    counts = defaultdict(int)
    hashtag_counter = Counter()
    mention_counter = Counter()
    keyword_counter = Counter()
    time_slot_counter = Counter()
    all_tweet_views = []

    for t in tweets:
        dt = t.created_at.astimezone(ist)
        text = unicodedata.normalize("NFKC", t.text.lower())
        
        counts["Total"] += 1
        slot = get_time_slot(dt)
        time_slot_counter[slot] += 1

        # Party-related classification
        for party, keywords in party_keywords.items():
            if party == "inc":
                if "sharmila" in text or "ys sharmila" in text:
                    if not any(tel_kw in text for tel_kw in telangana_keywords):
                        counts["INC_Related"] += 1
                continue
            if any(kw.lower() in text for kw in keywords):
                counts[f"{party.upper()}_Related"] += 1
                break

        # Govt keywords
        if any(gk in text for gk in govt_keywords):
            counts["Govt_Related"] += 1

        # Hashtags
        if t.entities and "hashtags" in t.entities:
            for tag in t.entities["hashtags"]:
                ht = "#" + tag["tag"].lower()
                hashtag_counter[ht] += 1

        # Mentions
        if t.entities and "mentions" in t.entities:
            for mention in t.entities["mentions"]:
                username = "@" + mention["username"].lower()
                mention_counter[username] += 1

        # Specific keywords
        for kw in specific_keywords:
            if kw.lower() in text:
                keyword_counter[kw] += 1

        # Tweet view info
        views = t.public_metrics.get("impression_count", 0)
        all_tweet_views.append({
            "views": views,
            "text": t.text,
            "url": f"https://x.com/{handle}/status/{t.id}"
        })

    print(f"✅ @{handle}: Total={counts['Total']} | TDP={counts['TDP_Related']} | YSRCP={counts['YSRCP_Related']} | JSP={counts['JSP_Related']} | BJP={counts['BJP_Related']} | INC={counts['INC_Related']} | Govt={counts['Govt_Related']}")

    # Top 3 most viewed tweets
    top3 = sorted(all_tweet_views, key=lambda x: x["views"], reverse=True)[:3]

    summary = {
        "Handle": handle,
        "Date": str(target_date),
        "Total Tweets": counts["Total"],
        "TDP Tweets": counts["TDP_Related"],
        "YSRCP Tweets": counts["YSRCP_Related"],
        "JSP Tweets": counts["JSP_Related"],
        "BJP Tweets": counts["BJP_Related"],
        "INC Tweets (Sharmila, AP only)": counts["INC_Related"],
        "Govt Related Tweets": counts["Govt_Related"],
        **{slot: time_slot_counter.get(slot, 0) for slot in time_slots},
        "Top 50 Hashtags": "; ".join(f"{ht}:{c}" for ht, c in hashtag_counter.most_common(50)),
        "Top 50 Mentions": "; ".join(f"{m}:{c}" for m, c in mention_counter.most_common(50)),
    }

    # Top 3 tweet details
    for i in range(3):
        if i < len(top3):
            summary[f"Top {i+1} Tweet Views"] = top3[i]["views"]
            summary[f"Top {i+1} Tweet URL"] = top3[i]["url"]
            summary[f"Top {i+1} Tweet Text"] = top3[i]["text"]
        else:
            summary[f"Top {i+1} Tweet Views"] = ""
            summary[f"Top {i+1} Tweet URL"] = ""
            summary[f"Top {i+1} Tweet Text"] = ""

    # Specific keyword mentions
    for kw in specific_keywords:
        summary[f"{kw}_mentions"] = keyword_counter.get(kw, 0)

    return summary
def run_in_batches(handles, batch_size=5):
    all_summaries = []
    for i in range(0, len(handles), batch_size):
        batch = handles[i:i + batch_size]
        with ThreadPoolExecutor(max_workers=len(batch)) as executor:
            futures = {executor.submit(process_handle, h): h for h in batch}
            for future in as_completed(futures):
                try:
                    result = future.result()
                    if result:
                        all_summaries.append(result)
                except Exception as e:
                    print(f"Error processing {futures[future]}: {e}")
        if i + batch_size < len(handles):
            time.sleep(90)
    return all_summaries
def format_and_send_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(12, max_len + 2)

    ws.freeze_panes = "A2"
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
        row[0].parent.row_dimensions[row[0].row].height = 22.5

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    wb.save(filename)

    yag = yagmail.SMTP(user=SENDER_EMAIL, password=SENDER_PASSWORD)
    subject = f"𝕏 Journalist Twitter Report – {dt.datetime.now().strftime('%d %B %Y')}"
    body = "Hi,\n\nPlease find below the attached journalist Twitter analysis report.\n\nRegards,\nNiveditha\nData Analyst\nShowtime Consulting"
    yag.send(to=TO_EMAIL, cc=CC_EMAIL, subject=subject, contents=body, attachments=[filename])

if __name__ == "__main__":
    output_filename = "journalist_twitter_analysis.xlsx"
    df = pd.DataFrame(run_in_batches(journalist_handles))
    df.to_excel(output_filename, index=False)
    format_and_send_excel(output_filename)
